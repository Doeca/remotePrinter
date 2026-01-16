#  本文件处理：付款信息变更
import os
import json
import dingLib
import shutil
import openpyxl
import pdfmod
import convert
import printmod
import handle
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from logsys import logger
alfont = Font(name="仿宋", size=12)
tifont = Font(name="仿宋", size=14, bold=True)
alstyle = Alignment(wrap_text=True, vertical='top')
images = list()


def copy_cell_style(source_cell, target_cell):
    """复制源单元格的字体、边框和填充样式到目标单元格"""
    if source_cell.has_style:
        # 复制字体
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                underline=source_cell.font.underline,
                color=source_cell.font.color
            )

        # 复制边框
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )

        # 复制填充
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
            


def get_value(raw: list, key: str, id: str = ''):
    val = None
    for v in raw:
        if v.name == key and id == '':
            val = v
        elif v.id == id:
            val = v
    if val != None:
        if val.component_type != "DDMultiSelectField":
            return val.value
        else:
            return "、".join(json.loads(s=val.value))
    return ""


def handle_xlsx(data, file):
    # 修改表格数据
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook["Sheet1"]
    worksheet['B2'].value = str(data.business_id)  # 审批编号
    worksheet['B3'].value = data.title.replace("提交的付款信息变更", "")  # 创建人
    worksheet['B4'].value = data.originator_dept_name  # 创建部门

    write_list = [
        ('B5', '申请人归属'),
        ('B6', '部门/门店'),
        ('B7', '企业主体'),
        ('B8', '是否已完成收款'),
        ('B9', '回款方式'),
        ('B10', '票种'),
        ('B11', '开票总额（元）'),
        ('B12', '收入类型'),
        ('B13', '收入类型时间段'),
        ('B14', '发票备注栏'),
        ('B15', '购方抬头'),
        ('B16', '购方纳税人识别号'),
        ('B17', '说明')
    ]
    for cell, key in write_list:
        worksheet[cell].value = get_value(data.form_component_values, key)

    goods_raw = get_value(data.form_component_values, '商品明细')
    if goods_raw != '':
        goods_detail = []
        try:
            for good in json.loads(s=goods_raw):
                row_value = ""
                row_values = good.get('rowValue', [])
                for obj in row_values:
                    label = obj.get('label', '')
                    value = obj.get('value', '')
                    row_value += f"{label}:{value} "
                goods_detail.append(row_value.strip())
            worksheet['B18'].value = "\n".join(goods_detail)  # 商品明细
        except Exception as e:
            logger.error(f"处理商品明细失败: {e}")
    st_id = 20
    raw_st_id = st_id
    # 获取源单元格的边框
    source_cell = worksheet[f'A{raw_st_id}']
    for vd in data.operation_records:
        line = ''
        if vd.type == "START_PROCESS_INSTANCE":
            line += f"{dingLib.getUserName(vd.user_id)} {vd.show_name}"
        elif vd.type == "EXECUTE_TASK_NORMAL":
            line += f"{dingLib.getUserName(vd.user_id)}"
            if vd.result == "AGREE":
                line += " 已同意 " + vd.date
            elif vd.result == "REJECT":
                line += " 已拒绝 " + vd.date
            elif vd.result == "CANCEL":
                line += " 已撤销 " + vd.date
        elif vd.type == "ADD_REMARK":
            line += f"{dingLib.getUserName(vd.user_id)} {vd.show_name}:{vd.remark}"
        elif vd.type == "PROCESS_CC":
            cc_ids = []
            for user in vd.cc_user_ids:
                cc_ids.append(dingLib.getUserName(user))
            line += "抄送给："+",".join(cc_ids)

        worksheet[f'A{st_id}'].alignment = alstyle
        worksheet[f'A{st_id}'].value = line
        copy_cell_style(source_cell, worksheet[f'A{st_id}'])
        worksheet.row_dimensions[st_id].height = worksheet.row_dimensions[raw_st_id].height
        worksheet.merge_cells(f'A{st_id}:B{st_id}')
        st_id += 1
    worksheet.print_area = f"A1:B{st_id-1}"  # 设置打印区域
    # 保存文件
    workbook.save(filename=file)


def handle_all(pid, title):
    global images
    images.clear()
    data = dingLib.getDetail(pid)
    shutil.copyfile("./template/kpsq.xlsx", f"./cache/{pid}_kpsq.xlsx")
    handle_xlsx(data, f"./cache/{pid}_kpsq.xlsx")
    convert.xlsx2pdf(f"./cache/{pid}_kpsq.xlsx", f"./cache/{pid}_temp.pdf")
    filename = f"./cache/{pid}_temp.pdf"
    for d in pdfmod.pdfexport(filename, f"{filename}_export"):
        images.append(d)

    # 将xlsx转成pdf，然后导出为图片，添加到images里
    # handle_attachments，收集整个表单中的附件，添加到images
    # handle.attachments(images, data, pid)
    # 把images merge成一个pdf
    pdfmod.mergeImages(f"./files/{pid}.pdf", images, 210, 297)
    # 递交打印
    printmod.a4_print(os.path.abspath(f"./files/{pid}.pdf"))
    logger.critical(f"{title} - {pid}.pdf 提交打印")
    return True


# handle_all('57iec76dQkOIMZkScFFiGA03401684753910')

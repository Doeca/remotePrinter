import os
import json
import dingLib
import shutil
import openpyxl
import pdfmod
import convert
import printmod
import handle
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from logsys import logger
alfont = Font(name="仿宋", size=12)
tifont = Font(name="仿宋", size=14, bold=True)
alstyle = Alignment(wrap_text=True, vertical='top')
images = list()


def copy_cell_style(source_cell, target_cell):
    """复制源单元格的字体、边框和填充样式到目标单元格"""
    if source_cell.has_style:
        # 复制字体
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                underline=source_cell.font.underline,
                color=source_cell.font.color
            )

        # 复制边框
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )

        # 复制填充
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )


def get_value(raw: list, key: str, id: str = ''):
    val = None
    for v in raw:
        if v.name == key and id == '':
            val = v
        elif v.id == id:
            val = v
    if val != None:
        if val.component_type != "DDMultiSelectField":
            return val.value
        else:
            return "、".join(json.loads(s=val.value))
    return ""


def handle_xlsx(data, file):
    # 修改表格数据
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook["Sheet1"]
    worksheet['B2'].value = str(data.business_id)  # 审批编号
    worksheet['B3'].value = data.title.replace("提交的付款信息变更", "")  # 创建人
    worksheet['B4'].value = data.originator_dept_name  # 创建部门

    write_list = [
        ('B5', '申请人归属'),
        ('B6', '部门/门店'),
        ('B7', '企业主体'),
        ('B8', '是否已完成收款'),
        ('B9', '回款方式'),
        ('B10', '票种'),

        ('B11', '红冲金额（负数）'),

        ('B13', '发票号码'),
        ('B14', '开票日期'),

        ('B16', '红冲发票原因'),

    ]
    for cell, key in write_list:
        worksheet[cell].value = get_value(data.form_component_values, key)

    raw_data = get_value(data.form_component_values, "原发票附件上传")
    fnames = []
    if raw_data != "":
        try:
            flist = json.loads(s=raw_data)
            if isinstance(flist, list):
                for vd in flist:
                    file_name = vd.get('fileName')
                    if file_name:
                        fnames.append(file_name)
        except Exception as e:
            logger.error(f"处理原发票附件失败: {e}")
    worksheet['B17'].value = "、".join(fnames)  # 原发票附件上传

    st_id = 19
    raw_st_id = st_id
    # 获取源单元格的边框
    source_cell = worksheet[f'A{raw_st_id}']
    for vd in data.operation_records:
        line = ''
        if vd.type == "START_PROCESS_INSTANCE":
            line += f"{dingLib.getUserName(vd.user_id)} {vd.show_name}"
        elif vd.type == "EXECUTE_TASK_NORMAL":
            line += f"{dingLib.getUserName(vd.user_id)}"
            if vd.result == "AGREE":
                line += " 已同意 " + vd.date
            elif vd.result == "REJECT":
                line += " 已拒绝 " + vd.date
            elif vd.result == "CANCEL":
                line += " 已撤销 " + vd.date
        elif vd.type == "ADD_REMARK":
            line += f"{dingLib.getUserName(vd.user_id)} {vd.show_name}:{vd.remark}"
        elif vd.type == "PROCESS_CC":
            cc_ids = []
            for user in vd.cc_user_ids:
                cc_ids.append(dingLib.getUserName(user))
            line += "抄送给："+",".join(cc_ids)

        worksheet[f'A{st_id}'].alignment = alstyle
        worksheet[f'A{st_id}'].value = line
        copy_cell_style(source_cell, worksheet[f'A{st_id}'])
        worksheet.row_dimensions[st_id].height = worksheet.row_dimensions[raw_st_id].height
        worksheet.merge_cells(f'A{st_id}:B{st_id}')
        st_id += 1
    worksheet.print_area = f"A1:B{st_id-1}"  # 设置打印区域
    # 保存文件
    workbook.save(filename=file)


def handle_all(pid, title):
    global images
    images.clear()
    data = dingLib.getDetail(pid)
    shutil.copyfile("./template/kphcsq.xlsx", f"./cache/{pid}_kphcsq.xlsx")
    handle_xlsx(data, f"./cache/{pid}_kphcsq.xlsx")
    convert.xlsx2pdf(f"./cache/{pid}_kphcsq.xlsx", f"./cache/{pid}_temp.pdf")
    filename = f"./cache/{pid}_temp.pdf"
    for d in pdfmod.pdfexport(filename, f"{filename}_export"):
        images.append(d)
    # 把images merge成一个pdf
    pdfmod.mergeImages(f"./files/{pid}.pdf", images, 210, 297)
    # 递交打印
    printmod.a4_print(os.path.abspath(f"./files/{pid}.pdf"))
    logger.critical(f"{title} - {pid}.pdf 提交打印")
    return True


# handle_all('57iec76dQkOIMZkScFFiGA03401684753910')

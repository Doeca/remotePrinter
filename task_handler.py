"""
統合タスクハンドラー
task_config.jsonの設定に基づいて各種タスクを処理
"""
import os
import json
import shutil
import time
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import dingLib
import pdfmod
import convert
import printmod
import handle
from logsys import logger

# フォント定義
alfont = Font(name="仿宋", size=12)
tifont = Font(name="仿宋", size=14, bold=True)
alstyle = Alignment(wrap_text=True, vertical='top')


class TaskHandler:
    """タスク処理の統合ハンドラー"""
    
    def __init__(self, config_file='template/task_config.json'):
        """初期化"""
        with open(config_file, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
    
    def get_value(self, raw_list, key, id=''):
        """フォームデータから値を取得"""
        val = None
        for v in raw_list:
            if v.name == key and id == '':
                val = v
            elif v.id == id:
                val = v
        
        if val is not None:
            if val.component_type != "DDMultiSelectField":
                return val.value
            else:
                return "、".join(json.loads(s=val.value))
        return ""
    
    def copy_cell_style(self, source_cell, target_cell):
        """セルスタイルをコピー"""
        if source_cell.has_style:
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    underline=source_cell.font.underline,
                    color=source_cell.font.color
                )
            if source_cell.border:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
    
    def format_goods_detail(self, raw_value):
        """商品明細のフォーマット"""
        if not raw_value:
            return ""
        
        try:
            goods_detail = []
            for good in json.loads(s=raw_value):
                row_value = ""
                row_values = good.get('rowValue', [])
                for obj in row_values:
                    label = obj.get('label', '')
                    value = obj.get('value', '')
                    row_value += f"{label}:{value} "
                goods_detail.append(row_value.strip())
            return "\n".join(goods_detail)
        except Exception as e:
            logger.error(f"处理商品明细失败: {e}")
            return ""
    
    def format_invoice_attachments(self, raw_value):
        """発票添付ファイルのフォーマット"""
        if not raw_value:
            return ""
        
        try:
            fnames = []
            flist = json.loads(s=raw_value)
            if isinstance(flist, list):
                for vd in flist:
                    file_name = vd.get('fileName')
                    if file_name:
                        fnames.append(file_name)
            return "、".join(fnames)
        except Exception as e:
            logger.error(f"处理原发票附件失败: {e}")
            return ""
    
    def process_task3_details(self, data, worksheet):
        """Task3専用の詳細処理（報銷明細）"""
        subFields = json.loads(s=self.get_value(data.form_component_values, "报销明细"))
        field_names = []
        field_moneys = []
        field_invoices = []
        
        for i in range(0, len(subFields)):
            row_values = subFields[i].get('rowValue', [])
            for subData in row_values:
                if subData.get('label', "") == '档口名称':
                    value = subData.get('value')
                    if value is not None:
                        field_names.append(str(value))
                if subData.get('label', "") == '报销金额（元）':
                    value = subData.get('value')
                    if value is not None:
                        field_moneys.append(str(value))
                if subData.get('label', "") == '发票':
                    rdata = subData.get('extendValue')
                    if not rdata:
                        continue
                    try:
                        if type(rdata) == str:
                            parsed_data = json.loads(s=rdata)
                            invoices = parsed_data.get('invoiceList', [])
                        else:
                            invoices = rdata.get('invoiceList', [])
                        for invoice in invoices:
                            invoice_no = invoice.get('invoiceNo')
                            if invoice_no:
                                field_invoices.append(invoice_no)
                    except Exception as e:
                        logger.error(f"处理发票数据失败: {e}")
        
        worksheet['B9'].value = "、".join(field_names)
        worksheet['B10'].value = "、".join(field_moneys)
        worksheet['B11'].value = "、".join(field_invoices)
        
        # 報銷総額処理
        total = self.get_value(data.form_component_values, "报销总额（元）")
        if total is not None:
            worksheet['B12'].value = total + "元"
        
        # 付款申請単と図片
        val_fksqd = self.get_value(data.form_component_values, "付款申请单")
        if val_fksqd is None:
            val_fksqd = ""
        worksheet['B13'].value = val_fksqd + self.get_value(data.form_component_values, "图片")
        worksheet['B14'].value = self.get_value(data.form_component_values, "附件")
    
    def fill_cell(self, worksheet, cell, mapping, data):
        """セルに値を入力"""
        mapping_type = mapping.get('type')
        
        if mapping_type == 'business_id':
            worksheet[cell].value = str(data.business_id)
        elif mapping_type == 'title_modified':
            config = self.config.get(str(data.task_type), {})
            title_replace = config.get('title_replace', '')
            worksheet[cell].value = data.title.replace(title_replace, "")
        elif mapping_type == 'originator_dept':
            worksheet[cell].value = data.originator_dept_name
        elif mapping_type == 'field':
            key = mapping.get('key')
            value = self.get_value(data.form_component_values, key)
            
            # フォーマット処理
            format_type = mapping.get('format')
            if format_type == 'goods_detail':
                value = self.format_goods_detail(value)
            elif format_type == 'invoice_attachments':
                value = self.format_invoice_attachments(value)
            
            worksheet[cell].value = value
    
    def add_operation_records(self, worksheet, data, config):
        """操作記録を追加"""
        start_row = config.get('operation_records_start', 16)
        style = config.get('operation_records_style', 'simple')
        
        st_id = start_row
        
        # タイトル追加（設定にある場合）
        if config.get('operation_records_title'):
            worksheet.merge_cells(f'A{st_id}:B{st_id}')
            worksheet[f'A{st_id}'].font = tifont
            worksheet[f'A{st_id}'].value = config.get('operation_records_title')
            st_id += 1
        
        # スタイルテンプレート取得
        if style == 'detailed':
            source_cell = worksheet[f'A{start_row}']
        
        for vd in data.operation_records:
            line = ''
            
            if style == 'simple':
                # シンプルスタイル（同意のみ）
                if vd.result != 'AGREE':
                    continue
                worksheet.merge_cells(f'A{st_id}:B{st_id}')
                if vd.remark != '':
                    worksheet[f'A{st_id}'].font = alfont
                    worksheet[f'A{st_id}'].alignment = alstyle
                    line = str(vd.remark) + "\n"
                line += dingLib.getUserName(vd.user_id) + "  已同意    " + vd.date
                worksheet[f'A{st_id}'].value = line
            else:
                # 詳細スタイル（全ての操作）
                if vd.type == "START_PROCESS_INSTANCE":
                    line = f"{dingLib.getUserName(vd.user_id)} {vd.show_name}"
                elif vd.type == "EXECUTE_TASK_NORMAL":
                    line = f"{dingLib.getUserName(vd.user_id)}"
                    if vd.result == "AGREE":
                        line += " 已同意 " + vd.date
                    elif vd.result == "REJECT":
                        line += " 已拒绝 " + vd.date
                    elif vd.result == "CANCEL":
                        line += " 已撤销 " + vd.date
                elif vd.type == "ADD_REMARK":
                    line = f"{dingLib.getUserName(vd.user_id)} {vd.show_name}:{vd.remark}"
                elif vd.type == "PROCESS_CC":
                    cc_ids = []
                    for user in vd.cc_user_ids:
                        cc_ids.append(dingLib.getUserName(user))
                    line = "抄送给：" + ",".join(cc_ids)
                
                worksheet[f'A{st_id}'].alignment = alstyle
                worksheet[f'A{st_id}'].value = line
                self.copy_cell_style(source_cell, worksheet[f'A{st_id}'])
                worksheet.row_dimensions[st_id].height = worksheet.row_dimensions[start_row].height
                worksheet.merge_cells(f'A{st_id}:B{st_id}')
            
            st_id += 1
        
        # 印刷エリア設定
        if config.get('print_area'):
            worksheet.print_area = f"A1:B{st_id-1}"
    
    def handle_xlsx(self, data, file, task_type):
        """Excelファイルを処理"""
        config = self.config.get(str(task_type))
        if not config:
            logger.error(f"未找到任务类型 {task_type} 的配置")
            return False
        
        # タスクタイプをデータに追加（title_modifiedで使用）
        data.task_type = task_type
        
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook["Sheet1"]
        
        # セルマッピング処理
        for cell, mapping in config.get('cell_mappings', {}).items():
            try:
                self.fill_cell(worksheet, cell, mapping, data)
            except Exception as e:
                logger.warning(f"处理单元格 {cell} 失败: {e}")
        
        # カスタム処理
        custom_processor = config.get('custom_processor')
        if custom_processor:
            processor_method = getattr(self, custom_processor, None)
            if processor_method:
                try:
                    processor_method(data, worksheet)
                except Exception as e:
                    logger.error(f"自定义处理失败: {e}")
        
        # 操作記録追加
        self.add_operation_records(worksheet, data, config)
        
        # 保存
        workbook.save(filename=file)
        return True
    
    def handle_all(self, pid, task_type, status='COMPLETED', title=''):
        """統合処理メイン関数"""
        config = self.config.get(str(task_type))
        if not config:
            logger.error(f"未找到任务类型 {task_type} 的配置")
            return False
        
        try:
            images = []
            data = dingLib.getDetail(pid)
            
            # Task3の条件分岐処理
            if task_type == 3 and status == 'RUNNING':
                continue_handle = False
                for vd in data.operation_records:
                    if vd.show_name == '店经理' and vd.result == 'AGREE':
                        continue_handle = True
                        break
                if not continue_handle:
                    logger.info("[Task3] 店经理暂未同意，不继续处理")
                    return False
                
                # 添付ファイル処理
                handle.attachments(images, data, pid)
                pdfmod.mergeImages(f"./files/{pid}_attachments.pdf", images, 210, 297)
                printmod.a4_print_singleside(os.path.abspath(f"./files/{pid}_attachments.pdf"))
                logger.critical(f"【附件】{title} - {pid}_attachments.pdf 提交打印")
                return True
            
            # 通常処理
            template = config.get('template')
            suffix = config.get('output_suffix', '')
            
            # テンプレートコピー
            template_base = os.path.splitext(template)[0]
            cache_file = f"./cache/{pid}_{template_base}.xlsx"
            
            # 既存ファイルを削除してからコピー（ロック回避）
            if os.path.exists(cache_file):
                try:
                    os.remove(cache_file)
                    time.sleep(0.1)  # 短い待機時間
                except PermissionError:
                    logger.warning(f"ファイルがロックされています、削除を再試行: {cache_file}")
                    time.sleep(1)
                    try:
                        os.remove(cache_file)
                    except Exception as e:
                        logger.error(f"ファイル削除失敗: {e}")
                        raise
            
            shutil.copyfile(f"./template/{template}", cache_file)
            
            # Excel処理
            self.handle_xlsx(data, cache_file, task_type)
            
            # PDF変換
            temp_pdf = f"./cache/{pid}_temp.pdf"
            convert.xlsx2pdf(cache_file, temp_pdf)
            
            # PDF→画像
            for d in pdfmod.pdfexport(temp_pdf, f"{temp_pdf}_export"):
                images.append(d)
            
            # 画像→PDF
            output_pdf = f"./files/{pid}{suffix}.pdf"
            pdfmod.mergeImages(output_pdf, images, 210, 297)
            
            # 印刷
            print_mode = config.get('print_mode', 'a4')
            if print_mode == 'a4':
                printmod.a4_print(os.path.abspath(output_pdf))
            elif print_mode == 'a4_singleside':
                printmod.a4_print_singleside(os.path.abspath(output_pdf))
            
            logger.critical(f"{title} - {os.path.basename(output_pdf)} 提交打印")
            return True
            
        except Exception as e:
            logger.exception(f"处理任务失败 - PID: {pid}, Type: {task_type}")
            return False


# グローバルハンドラーインスタンス
_handler = None

def get_handler():
    """ハンドラーインスタンスを取得（シングルトン）"""
    global _handler
    if _handler is None:
        _handler = TaskHandler()
    return _handler


def handle_task(pid, task_type, status='COMPLETED', title=''):
    """タスク処理のエントリーポイント"""
    handler = get_handler()
    return handler.handle_all(pid, task_type, status, title)

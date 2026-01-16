#  本文件处理：付款信息变更
import os
import json
import dingLib
import shutil
import openpyxl
import pdfmod
import convert
import printmod
import handle
from openpyxl.styles import Alignment, Font
from logsys import logger
alfont = Font(name="仿宋", size=12)
tifont = Font(name="仿宋", size=14, bold=True)
alstyle = Alignment(wrap_text=True, vertical='top')
images = list()


def get_value(raw: list, key: str, id: str = ''):
    val = None
    for v in raw:
        if v.name == key and id == '':
            val = v
        elif v.id == id:
            val = v
    if val != None:
        if val.component_type != "DDMultiSelectField":
            return val.value
        else:
            return "、".join(json.loads(s=val.value))
    return ""


def handle_xlsx(data, file):
    # 修改表格数据
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook["Sheet1"]
    worksheet['B2'].value = str(data.business_id)  # 审批编号
    worksheet['B3'].value = data.title.replace("提交的付款信息变更", "")  # 创建人
    worksheet['B4'].value = data.originator_dept_name  # 创建部门
    worksheet['B5'].value = get_value(
        data.form_component_values, "申请门店")
    worksheet['B6'].value = get_value(
        data.form_component_values, "档口名称")
    worksheet['B7'].value = get_value(
        data.form_component_values, "信息错误类型")
    worksheet['B8'].value = get_value(
        data.form_component_values, "其他补充")
    worksheet['B10'].value = get_value(
        data.form_component_values, "户名")
    worksheet['B11'].value = get_value(
        data.form_component_values, "银行卡号")
    worksheet['B12'].value = get_value(
        data.form_component_values, "银行开户行")
    worksheet['B13'].value = get_value(
        data.form_component_values, "银行卡照片")
    worksheet['B15'].value = get_value(
        data.form_component_values, "备注")

    # 从16开始自行添加
    st_id = 16
    # 审批流程标题
    worksheet.merge_cells(f'A{st_id}:B{st_id}')
    worksheet[f'A{st_id}'].font = tifont
    worksheet[f'A{st_id}'].value = "审批流程"
    st_id += 1
    for vd in data.operation_records:
        if vd.result != 'AGREE':
            continue
        worksheet.merge_cells(f'A{st_id}:B{st_id}')
        line = ''
        if (vd.remark != ''):
            worksheet[f'A{st_id}'].alignment = alstyle
            line = str(vd.remark) + "\n"
        line += dingLib.getUserName(vd.user_id) + "  已同意    "+vd.date
        worksheet[f'A{st_id}'].value = line
        st_id += 1

    # 保存文件
    workbook.save(filename=file)


def handle_all(pid, title):
    global images
    images.clear()
    data = dingLib.getDetail(pid)
    shutil.copyfile("./template/fkxxbg.xlsx", f"./cache/{pid}_fkxxbg.xlsx")
    handle_xlsx(data, f"./cache/{pid}_fkxxbg.xlsx")
    convert.xlsx2pdf(f"./cache/{pid}_fkxxbg.xlsx", f"./cache/{pid}_temp.pdf")
    filename = f"./cache/{pid}_temp.pdf"
    for d in pdfmod.pdfexport(filename, f"{filename}_export"):
        images.append(d)

    # 将xlsx转成pdf，然后导出为图片，添加到images里
    # handle_attachments，收集整个表单中的附件，添加到images
    # handle.attachments(images, data, pid)
    # 把images merge成一个pdf
    pdfmod.mergeImages(f"./files/{pid}.pdf", images, 210, 297)
    # 递交打印
    printmod.a4_print(os.path.abspath(f"./files/{pid}.pdf"))
    logger.critical(f"{title} - {pid}.pdf 提交打印")
    return True


# handle_all('57iec76dQkOIMZkScFFiGA03401684753910')
